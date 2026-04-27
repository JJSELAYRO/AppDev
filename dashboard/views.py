from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count
from django.http import HttpResponse
from accounts.models import UserProfile, Teacher, Student, Parent, ClassGroup
from attendance.models import Attendance, QRSession, SMSLog
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io


def dashboard_redirect(request):
    """Redirect users to their appropriate dashboard based on role"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        profile = request.user.profile
        if profile.role == 'admin':
            return redirect('admin_dashboard')
        elif profile.role == 'teacher':
            return redirect('teacher_dashboard')
        elif profile.role == 'student':
            return redirect('student_dashboard')
        elif profile.role == 'parent':
            return redirect('parent_dashboard')
    except UserProfile.DoesNotExist:
        pass
    
    return redirect('login')


@login_required
def admin_dashboard(request):
    """Admin dashboard with user management and reports"""
    try:
        if request.user.profile.role != 'admin':
            messages.error(request, 'Access denied. Admin only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # Statistics
    total_students = Student.objects.count()
    total_teachers = Teacher.objects.count()
    total_parents = Parent.objects.count()
    total_classes = ClassGroup.objects.count()
    
    # Today's attendance
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(date=today).count()
    
    # Recent attendances
    recent_attendances = Attendance.objects.select_related(
        'student', 'student__user_profile__user', 'class_group'
    ).order_by('-created_at')[:10]
    
    # SMS logs
    recent_sms = SMSLog.objects.select_related('parent').order_by('-timestamp')[:10]
    
    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_parents': total_parents,
        'total_classes': total_classes,
        'today_attendance': today_attendance,
        'recent_attendances': recent_attendances,
        'recent_sms': recent_sms,
    }
    return render(request, 'dashboard/admin_dashboard_new.html', context)


@login_required
def teacher_dashboard(request):
    """Teacher dashboard for creating sessions and viewing attendance"""
    try:
        if request.user.profile.role != 'teacher':
            messages.error(request, 'Access denied. Teacher only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    teacher = request.user.profile.teacher_profile
    
    # Active sessions
    active_sessions = QRSession.objects.filter(
        teacher=teacher,
        is_active=True,
        expires_at__gt=timezone.now()
    ).select_related('class_group').order_by('-created_at')
    
    # Advised classes
    advised_classes = ClassGroup.objects.filter(adviser=teacher)
    
    # Today's attendance for teacher's classes
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(
        class_group__in=advised_classes,
        date=today
    ).select_related('student', 'student__user_profile__user').order_by('-time_in')
    
    # Attendance statistics
    attendance_stats = Attendance.objects.filter(
        class_group__in=advised_classes,
        date=today
    ).values('status').annotate(count=Count('id'))
    
    context = {
        'teacher': teacher,
        'active_sessions': active_sessions,
        'advised_classes': advised_classes,
        'today_attendance': today_attendance,
        'attendance_stats': {stat['status']: stat['count'] for stat in attendance_stats},
    }
    return render(request, 'dashboard/teacher_dashboard_new.html', context)


@login_required
def student_dashboard(request):
    """Student dashboard with QR scanner and attendance history"""
    try:
        if request.user.profile.role != 'student':
            messages.error(request, 'Access denied. Student only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    student = request.user.profile.student_profile
    
    # Attendance history
    attendance_history = Attendance.objects.filter(
        student=student
    ).select_related('class_group', 'session').order_by('-date', '-time_in')[:20]
    
    # Attendance statistics
    total_attendance = Attendance.objects.filter(student=student).count()
    present_count = Attendance.objects.filter(student=student, status='present').count()
    late_count = Attendance.objects.filter(student=student, status='late').count()
    absent_count = Attendance.objects.filter(student=student, status='absent').count()
    
    context = {
        'student': student,
        'attendance_history': attendance_history,
        'total_attendance': total_attendance,
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
    }
    return render(request, 'dashboard/student_dashboard_new.html', context)


@login_required
def parent_dashboard(request):
    """Parent dashboard to view children's attendance"""
    try:
        if request.user.profile.role != 'parent':
            messages.error(request, 'Access denied. Parent only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    parent = request.user.profile.parent_profile
    
    # Get children
    children = Student.objects.filter(parent=parent).select_related(
        'user_profile__user', 'class_group'
    )
    
    # Get attendance for all children
    children_attendance = []
    for child in children:
        attendance = Attendance.objects.filter(
            student=child
        ).select_related('class_group').order_by('-date', '-time_in')[:10]
        children_attendance.append({
            'child': child,
            'attendance': attendance,
            'total_attendance': Attendance.objects.filter(student=child).count(),
            'present_count': Attendance.objects.filter(student=child, status='present').count(),
        })
    
    # SMS logs
    sms_logs = SMSLog.objects.filter(parent=parent).order_by('-timestamp')[:20]
    
    context = {
        'parent': parent,
        'children_attendance': children_attendance,
        'sms_logs': sms_logs,
    }
    return render(request, 'dashboard/parent_dashboard.html', context)


@login_required
def upload_students_excel(request):
    """Admin view to upload Excel file and create student accounts"""
    try:
        if request.user.profile.role != 'admin':
            messages.error(request, 'Access denied. Admin only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # Read Excel file
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            error_count = 0
            
            # Skip header row, start from row 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                try:
                    student_id = str(row[0])
                    first_name = str(row[1])
                    last_name = str(row[2])
                    email = str(row[3]) if row[3] else f"{student_id}@student.school.edu"
                    class_name = str(row[4]) if len(row) > 4 and row[4] else None
                    parent_name = str(row[5]) if len(row) > 5 and row[5] else None
                    parent_phone = str(row[6]) if len(row) > 6 and row[6] else None
                    
                    # Create user with default password (student ID)
                    username = student_id.lower()
                    default_password = student_id  # Default password is student ID
                    
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            'email': email,
                            'first_name': first_name,
                            'last_name': last_name,
                        }
                    )
                    
                    if created:
                        user.set_password(default_password)
                        user.save()
                        
                        # Update profile
                        user.profile.role = 'student'
                        user.profile.must_change_password = True  # Flag to force password change
                        user.profile.save()
                        
                        # Get or create class
                        class_group = None
                        if class_name:
                            class_group, _ = ClassGroup.objects.get_or_create(name=class_name)
                        
                        # Get or create parent
                        parent = None
                        if parent_name and parent_phone:
                            parent_user, _ = User.objects.get_or_create(
                                username=f"parent_{student_id}",
                                defaults={
                                    'email': f"parent_{student_id}@school.edu",
                                    'first_name': parent_name,
                                }
                            )
                            if _:
                                parent_user.set_password(default_password)
                                parent_user.save()
                                parent_user.profile.role = 'parent'
                                parent_user.profile.save()
                                
                                parent, _ = Parent.objects.get_or_create(
                                    user_profile=parent_user.profile,
                                    defaults={'contact_number': parent_phone}
                                )
                        
                        # Create student profile
                        Student.objects.get_or_create(
                            user_profile=user.profile,
                            defaults={
                                'student_id': student_id,
                                'class_group': class_group,
                                'parent': parent
                            }
                        )
                        
                        created_count += 1
                    
                except Exception as e:
                    error_count += 1
                    print(f"Error creating student: {e}")
                    continue
            
            if created_count > 0:
                messages.success(request, f'Successfully created {created_count} student accounts! Default password is their Student ID.')
            if error_count > 0:
                messages.warning(request, f'{error_count} rows had errors and were skipped.')
            
            return redirect('admin_dashboard')
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
            return redirect('admin_dashboard')
    
    return redirect('admin_dashboard')


@login_required
def download_students_template(request):
    """Download Excel template for student upload"""
    try:
        if request.user.profile.role != 'admin':
            messages.error(request, 'Access denied. Admin only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Headers
    headers = ['Student ID', 'First Name', 'Last Name', 'Email', 'Class', 'Parent Name', 'Parent Phone']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="2D6CDF", end_color="2D6CDF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample data
    ws.append(['STU001', 'Juan', 'Dela Cruz', 'juan@school.edu', 'Grade 10-A', 'Maria Dela Cruz', '+639123456789'])
    ws.append(['STU002', 'Maria', 'Santos', 'maria@school.edu', 'Grade 10-A', 'Jose Santos', '+639987654321'])
    
    # Adjust column widths
    column_widths = [15, 15, 15, 25, 15, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=student_template.xlsx'
    
    wb.save(response)
    return response


@login_required
def download_all_students(request):
    """Download all students as Excel"""
    try:
        if request.user.profile.role != 'admin':
            messages.error(request, 'Access denied. Admin only.')
            return redirect('dashboard')
    except:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Students"
    
    # Headers
    headers = ['Student ID', 'Name', 'Email', 'Class', 'Default Password', 'Must Change Password']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="2D6CDF", end_color="2D6CDF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add student data
    students = Student.objects.select_related('user_profile__user', 'class_group').all()
    for student in students:
        ws.append([
            student.student_id,
            f"{student.user_profile.user.first_name} {student.user_profile.user.last_name}",
            student.user_profile.user.email,
            student.class_group.name if student.class_group else 'Not assigned',
            student.student_id,  # Default password
            'Yes' if student.user_profile.user.profile.must_change_password else 'No'
        ])
    
    # Adjust column widths
    column_widths = [15, 25, 25, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=all_students.xlsx'
    
    wb.save(response)
    return response
